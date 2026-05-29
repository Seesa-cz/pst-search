"""Email indexer — parses .eml files, extracts text + attachments, syncs to EmailStore."""

from __future__ import annotations
import email
import email.header
import hashlib
import logging
import time
from email.utils import parsedate_to_datetime
from pathlib import Path

from .store import EmailStore

logger = logging.getLogger(__name__)

_ENCODE_BATCH = 200     # embeddings per fastembed call
_PROGRESS_INTERVAL = 15.0  # seconds between progress log lines


def email_id(path: Path, base: Path) -> str:
    return hashlib.sha256(str(path.relative_to(base)).encode()).hexdigest()


def chunk_id(eid: str, idx: int) -> str:
    return f"{eid}:{idx}"


# ---------------------------------------------------------------------------
# Header / body helpers
# ---------------------------------------------------------------------------

def _decode_header(value: str | None) -> str:
    if not value:
        return ""
    parts = email.header.decode_header(value)
    result = []
    for raw, charset in parts:
        if isinstance(raw, bytes):
            try:
                result.append(raw.decode(charset or "utf-8", errors="replace"))
            except Exception:
                result.append(raw.decode("latin-1", errors="replace"))
        else:
            result.append(str(raw))
    return " ".join(result).strip()


def _get_body(msg: email.message.Message) -> str:
    parts = []
    for part in msg.walk():
        ct = part.get_content_type()
        disp = str(part.get("Content-Disposition", ""))
        if ct == "text/plain" and "attachment" not in disp:
            payload = part.get_payload(decode=True)
            if payload:
                charset = part.get_content_charset() or "utf-8"
                try:
                    parts.append(payload.decode(charset, errors="replace"))
                except Exception:
                    parts.append(payload.decode("latin-1", errors="replace"))
    return "\n".join(parts).strip()


# ---------------------------------------------------------------------------
# Attachment text extraction
# ---------------------------------------------------------------------------

def _extract_pdf(data: bytes) -> str:
    try:
        import io
        from pdfminer.high_level import extract_text
        return (extract_text(io.BytesIO(data)) or "")[:8000]
    except Exception:
        return ""


def _extract_docx(data: bytes) -> str:
    try:
        import io
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text)[:8000]
    except Exception:
        return ""


def _extract_xlsx(data: bytes) -> str:
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(" ".join(cells))
            if len(parts) > 500:
                break
        return "\n".join(parts)[:8000]
    except Exception:
        return ""


def _get_attachments(msg: email.message.Message) -> list[tuple[str, str]]:
    """Returns [(filename, extracted_text)] for each attachment."""
    results = []
    for part in msg.walk():
        disp = str(part.get("Content-Disposition", ""))
        if "attachment" not in disp:
            continue
        filename = _decode_header(part.get_filename() or "")
        if not filename:
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            results.append((filename, ""))
            continue
        fn_lower = filename.lower()
        if fn_lower.endswith(".pdf"):
            text = _extract_pdf(payload)
        elif fn_lower.endswith(".docx"):
            text = _extract_docx(payload)
        elif fn_lower.endswith(".xlsx") or fn_lower.endswith(".xls"):
            text = _extract_xlsx(payload)
        else:
            text = ""
        results.append((filename, text))
    return results


# ---------------------------------------------------------------------------
# .eml → chunks
# ---------------------------------------------------------------------------

def parse_eml(path: Path, base: Path) -> list[dict]:
    """Parse a single .eml file; return list of chunk dicts for the store."""
    try:
        with open(path, "rb") as f:
            msg = email.message_from_binary_file(f)
    except Exception as e:
        logger.warning("Cannot parse %s: %s", path, e)
        return []

    subject = _decode_header(msg.get("Subject", "(no subject)"))
    from_addr = _decode_header(msg.get("From", ""))
    to_addr = _decode_header(msg.get("To", ""))
    cc_addr = _decode_header(msg.get("Cc", ""))

    date_str = ""
    try:
        date_raw = msg.get("Date", "")
        if date_raw:
            dt = parsedate_to_datetime(date_raw)
            date_str = dt.isoformat()
    except Exception:
        date_str = msg.get("Date", "") or ""

    folder = str(path.parent.relative_to(base))
    body = _get_body(msg)
    attachments = _get_attachments(msg)

    eid = email_id(path, base)
    rel_path = str(path.relative_to(base))
    att_names = ", ".join(fn for fn, _ in attachments)
    has_att = "true" if attachments else "false"

    base_meta = {
        "subject":          subject,
        "from_addr":        from_addr,
        "to_addr":          to_addr,
        "cc_addr":          cc_addr,
        "date_str":         date_str,
        "folder":           folder,
        "has_attachment":   has_att,
        "attachment_names": att_names,
        "path":             rel_path,
        "file_id":          eid,
    }

    chunks = []

    # Body chunk — always included
    body_embed = (
        f"Subject: {subject}\nFrom: {from_addr}\nTo: {to_addr}"
        f"\nDate: {date_str}\nFolder: {folder}\n\n{body}"
    )
    chunks.append({
        "id":       chunk_id(eid, 0),
        "document": body_embed,
        "metadata": {**base_meta, "chunk_heading": "body", "chunk_idx": 0},
    })

    # Attachment chunks — only when text is extractable
    for idx, (filename, text) in enumerate(attachments, 1):
        if not text.strip():
            continue
        att_embed = (
            f"Subject: {subject}\nFrom: {from_addr}"
            f"\nAttachment: {filename}\n\n{text}"
        )
        chunks.append({
            "id":       chunk_id(eid, idx),
            "document": att_embed,
            "metadata": {**base_meta, "chunk_heading": f"attachment: {filename}", "chunk_idx": idx},
        })

    return chunks


# ---------------------------------------------------------------------------
# Sync
# ---------------------------------------------------------------------------

def sync(eml_dir: Path, store: EmailStore) -> int:
    """Scan eml_dir for .eml files and index any that are new or modified.

    Returns number of emails (re)indexed.
    """
    eml_files = list(eml_dir.rglob("*.eml"))
    if not eml_files:
        logger.warning("No .eml files found in %s", eml_dir)
        return 0

    known = store.get_all_file_ids()
    to_index = [p for p in eml_files if email_id(p, eml_dir) not in known]

    total = len(to_index)
    if total == 0:
        logger.warning("All %d emails already indexed", len(eml_files))
        return 0

    logger.warning("Indexing %d new emails (total on disk: %d)", total, len(eml_files))

    all_ids, all_docs, all_metas, all_texts = [], [], [], []
    indexed = 0
    t_last = time.monotonic()

    for path in to_index:
        chunks = parse_eml(path, eml_dir)
        for c in chunks:
            all_ids.append(c["id"])
            all_docs.append(c["document"])
            all_metas.append(c["metadata"])
            all_texts.append(c["document"])

        # Flush when batch is full
        if len(all_ids) >= _ENCODE_BATCH:
            embeddings = store.encode_batch(all_texts)
            store.upsert_batch(all_ids, embeddings, all_docs, all_metas)
            indexed += sum(1 for m in all_metas if m["chunk_idx"] == 0)
            all_ids, all_docs, all_metas, all_texts = [], [], [], []

        now = time.monotonic()
        if now - t_last >= _PROGRESS_INTERVAL:
            logger.warning("Indexing: %d/%d emails processed", indexed, total)
            t_last = now

    # Final flush
    if all_ids:
        embeddings = store.encode_batch(all_texts)
        store.upsert_batch(all_ids, embeddings, all_docs, all_metas)
        indexed += sum(1 for m in all_metas if m["chunk_idx"] == 0)

    logger.warning("Indexed %d emails, %d total chunks", indexed, store.count())
    return indexed


def full_reindex(eml_dir: Path, store: EmailStore) -> int:
    """Drop all existing entries and re-index everything from scratch."""
    data_dir = store._data_dir
    import sqlite3
    if store._db:
        store._db.close()
        store._db = None
    store._mat = None
    store._row_ids = []
    store._rid_to_idx = {}
    store._vectors_loaded = False
    store._bm25 = None
    store._bm25_dirty = True
    for f in ("vectors.npy", "row_ids.npy", "metadata.db"):
        (data_dir / f).unlink(missing_ok=True)
    return sync(eml_dir, store)
